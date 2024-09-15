from flask import Flask, request, jsonify, send_file, abort
import os
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment
from openpyxl.drawing.image import Image
import openai
import matplotlib.pyplot as plt
from pandas.tseries.frequencies import to_offset
from matplotlib.dates import MinuteLocator, SecondLocator, DateFormatter
from matplotlib.ticker import MaxNLocator
import io

app = Flask(__name__)

# ตั้งค่า OpenAI API key ของคุณ
openai.api_key = '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'  # เปลี่ยนเป็น API key ที่ถูกต้อง

# กำหนดที่เก็บไฟล์
UPLOAD_FOLDER = 'uploads'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# ชื่อไฟล์ที่ต้องการดาวน์โหลด
DOWNLOAD_FILE_NAME = 'result_with_summary.xlsx'

# ฟังก์ชันสร้างและส่งออกกราฟ
def create_graph_and_export(df, interval, output_path):
    resampled_data = df['elapsed'].resample(interval).mean()  # Resample ข้อมูลตามช่วงเวลาที่คำนวณได้
    resampled_data = resampled_data.fillna(0)  # เติม NaN ด้วย 0 หรือค่าอื่นที่เหมาะสม

    max_value = resampled_data.max()  # หาค่าจุดสูงสุด (Maximum)
    max_time = resampled_data.idxmax()

    # ดึงข้อมูลช่วงเวลาระหว่าง timestamps เริ่มต้นและสิ้นสุด
    start_time = df.index.min()
    end_time = df.index.max()
    total_duration = end_time - start_time

    # แปลง total_duration เป็นข้อความที่เข้าใจง่าย
    if total_duration.days > 0:
        duration_str = f"{total_duration.days} day(s) {total_duration.seconds // 3600} hr(s) {total_duration.seconds % 3600 // 60} m"
    elif total_duration.seconds >= 3600:
        duration_str = f"{total_duration.seconds // 3600} hr(s) {total_duration.seconds % 3600 // 60} m"
    elif total_duration.seconds >= 60:
        duration_str = f"{total_duration.seconds // 60} m {total_duration.seconds % 60} s"
    else:
        duration_str = f"{total_duration.seconds} s"

    # สร้างกราฟเส้นพร้อม markers
    fig, ax = plt.subplots(figsize=(14, 8))  # ปรับขนาด figure ให้เหมาะสมตามข้อมูลที่ต้องการแสดงผล
    ax.plot(resampled_data.index, resampled_data.values, marker='o', linestyle='-', color='b')
    ax.scatter(max_time, max_value, color='r', zorder=5)  # เพิ่มจุดสีแดงสำหรับจุดสูงสุด
    ax.annotate(f'Max: {max_value:.2f}',
                 xy=(max_time, max_value),
                 xytext=(max_time, max_value + 5),
                 arrowprops=dict(facecolor='red', shrink=0.05),
                 fontsize=12,
                 ha='center')

    # ตั้งค่ารูปแบบของ x-axis เพื่อแสดงผลตามช่วงเวลาที่คำนวณได้
    ax.set_title(f'Average Response Time : {duration_str}', fontsize=16)
    ax.set_xlabel('Timestamp', fontsize=14)
    ax.set_ylabel('Average Elapsed Time (ms)', fontsize=14)
    ax.grid(True)

    interval_seconds = int(to_offset(interval).nanos / 1e9)

    # เลือกใช้ SecondLocator หรือ MinuteLocator ตามค่า interval ที่คำนวณได้
    if interval_seconds < 60:
        locator = SecondLocator(bysecond=range(0, 60, max(1, interval_seconds)))
    else:
        locator = MinuteLocator(byminute=range(0, 60, max(1, interval_seconds // 60)))

    ax.xaxis.set_major_locator(locator)
    formatter = DateFormatter('%H:%M:%S' if 'S' in interval else '%H:%M')
    ax.xaxis.set_major_formatter(formatter)

    # ลดจำนวน ticks ในแกน x เพื่อไม่ให้ตัวเลขของแกน x มากเกินไป
    ax.xaxis.set_major_locator(MaxNLocator(nbins=10))

    fig.autofmt_xdate()  # Auto-fix for overlapping dates

    # หมุนตัวเลขในแกน x ให้เป็นแนวตั้ง
    plt.xticks(rotation=90)

    fig.tight_layout()  # ใช้ tight_layout เพื่ออัปเดตการแสดงผลอย่างเหมาะสม

    # บันทึกกราฟเป็น PNG
    graph_path = "graph.png"
    fig.savefig(graph_path)

    # อ่านเทมเพลต Excel
    workbook = openpyxl.load_workbook(output_path)
    sheet3 = workbook.create_sheet("Sheet3")

    # แทรกรูปเข้าที่ Sheet3
    img = Image(graph_path)
    sheet3.add_image(img, 'A1')

    # บันทึกผลลัพธ์ลงไฟล์ Excel ที่กำหนด
    workbook.save(output_path)
    print(f"Graph added to {output_path}")

# ฟังก์ชันประมวลผลไฟล์
def process_file(file_path, interval):
    try:
        data = pd.read_csv(file_path)
    except Exception as e:
        print(f"Error: Unable to read CSV file. {e}")
        abort(400, description="Unable to read CSV file.")

    # ตรวจสอบคอลัมน์ที่มีอยู่
    print("Columns in the DataFrame:", data.columns)

    # ตรวจสอบว่ามีคอลัมน์ 'label' อยู่ใน DataFrame หรือไม่
    label_column = 'label'
    if label_column not in data.columns:
        print("Error: 'label' column not found in the data!")
        abort(400, description="'label' column not found in the data!")

    # ตรวจสอบค่าในคอลัมน์ 'success'
    print("Unique values in 'success' column:", data['success'].unique())

    # แปลงค่า `elapsed` จาก ms เป็น s
    data['elapsed_seconds'] = data['elapsed'] / 1000

    # คำนวณค่า Pass, Fail และ %Error Rate
    data['Pass'] = data['success'].apply(lambda x: 1 if x else 0)
    data['Fail'] = data['success'].apply(lambda x: 0 if x else 1)

    # รวมค่าเส้น API ที่มีชื่อเดียวกัน
    summarized_data = data.groupby('label').agg(
        Transaction_Name=('label', 'first'),
        Expected=('label', 'size'),
        Pass=('Pass', 'sum'),
        Fail=('Fail', 'sum'),
        AvgResponseTime=('elapsed_seconds', 'mean'),
        MinResponseTime=('elapsed_seconds', 'min'),
        MaxResponseTime=('elapsed_seconds', 'max')
    ).reset_index(drop=True)

    # คำนวณ %Error Rate หลังจากการรวมข้อมูล
    summarized_data['%Error Rate'] = (summarized_data['Fail'] / summarized_data['Expected']) * 100

    # กำหนดค่า SLA ของ Response Time ให้เป็น 3 ทั้งหมด
    summarized_data['SLA_ResponseTime'] = 3

    # สร้างคอลัมน์ Status สำหรับการตรวจสอบการผ่านหรือไม่ผ่าน
    summarized_data['Status'] = summarized_data['%Error Rate'].apply(lambda x: "Pass" if x < 2 else "Failed")

    # โหลดเทมเพลต Excel
    template_path = "template.xlsx"
    workbook = openpyxl.load_workbook(template_path)
    sheet2 = workbook['Sheet2']

    # กำหนดสีสำหรับสถานะ
    green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

    # กำหนดกรอบรอบบรรทัดที่มีข้อมูล
    thin_border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))

    # กำหนดการจัดตำแหน่งชิดขวาสำหรับข้อมูลตัวเลข
    right_align = Alignment(horizontal='right')

    # เขียนข้อมูลที่รวมค่าแล้วลง Excel (Sheet2)
    for index, row in summarized_data.iterrows():
        start_row = index + 3  # สามารถปรับแถวเริ่มต้นตามที่ต้องการ

        sheet2.cell(row=start_row, column=1, value=row['Transaction_Name']).border = thin_border

        cell_expected = sheet2.cell(row=start_row, column=2, value=row['Expected'])
        cell_expected.border = thin_border
        cell_expected.alignment = right_align

        cell_pass = sheet2.cell(row=start_row, column=3, value=row['Pass'])
        cell_pass.border = thin_border
        cell_pass.alignment = right_align

        cell_fail = sheet2.cell(row=start_row, column=4, value=row['Fail'])
        cell_fail.border = thin_border
        cell_fail.alignment = right_align

        cell_error_rate = sheet2.cell(row=start_row, column=5, value=f"{row['%Error Rate']:.2f}%")
        cell_error_rate.border = thin_border
        cell_error_rate.alignment = right_align

        status_cell = sheet2.cell(row=start_row, column=6, value=row['Status'])
        if row['%Error Rate'] < 2:
            status_cell.fill = green_fill
        else:
            status_cell.fill = red_fill
        status_cell.border = thin_border

        cell_sla_response_time = sheet2.cell(row=start_row, column=7, value=row['SLA_ResponseTime'])
        cell_sla_response_time.border = thin_border
        cell_sla_response_time.alignment = right_align

        cell_avg_response_time = sheet2.cell(row=start_row, column=8, value=row['AvgResponseTime'])
        cell_avg_response_time.border = thin_border
        cell_avg_response_time.alignment = right_align

        cell_min_response_time = sheet2.cell(row=start_row, column=9, value=row['MinResponseTime'])
        cell_min_response_time.border = thin_border
        cell_min_response_time.alignment = right_align

        cell_max_response_time = sheet2.cell(row=start_row, column=10, value=row['MaxResponseTime'])
        cell_max_response_time.border = thin_border
        cell_max_response_time.alignment = right_align

        # เปรียบเทียบ SLA ของ Response Time กับ Average
        sla_status_cell = sheet2.cell(row=start_row, column=11)
        if row['AvgResponseTime'] <= row['SLA_ResponseTime']:
            sla_status_cell.value = "Pass"
            sla_status_cell.fill = green_fill
        else:
            sla_status_cell.value = "Failed"
            sla_status_cell.fill = red_fill
        sla_status_cell.border = thin_border
        sla_status_cell.alignment = right_align

    # **การสรุปผลด้วย OpenAI API**

    # อ่านข้อมูลจาก Sheet2 เข้า DataFrame อีกครั้ง
    sheet2_data = pd.DataFrame(sheet2.values)
    sheet2_data.columns = sheet2_data.iloc[0]  # ตั้งค่าแถวที่ 0 เป็นชื่อคอลัมน์
    sheet2_data = sheet2_data[1:]  # เอาข้อมูลจริงๆ ที่เริ่มจากแถวที่ 1 เป็นต้นไป

    # รวมข้อมูลเพิ่มเติมที่ต้องการ
    test_type = "Load Test"  # ตัวอย่างของประเภทที่ใช้ในการทดสอบ
    total_users = data['threadName'].nunique()  # จำนวนนักทดสอบ (user) ทั้งหมด
    project_name = "Example Project"  # ตัวอย่างของชื่อโปรเจค

    # แปลง DataFrame เป็นข้อความเพื่อส่งไปยัง OpenAI
    text_data = sheet2_data.to_string(index=False)

    # โปรมต์ที่ต้องการให้ OpenAI สรุปข้อมูล
    prompt = (f"Please summarize the following data:\n\n{text_data}\n\n"
            f"Test Type: {test_type}\n"
            f"Total Users: {total_users}\n"
            f"Project Name: {project_name}\n\n"
            f"Detail:\n"
            f"- Show all Passed , SLA transactions\n"
            f"- Show all Failed , SLA transactions\n\n"
            f"Summary:")

    # เรียกใช้ OpenAI API เพื่อให้สรุปข้อมูล
    response = openai.ChatCompletion.create(
        model="gpt-4",
        messages=[
            {"role": "system", "content": "You are a data analyst."},
            {"role": "user", "content": prompt}
        ]
    )

    summary = response.choices[0].message['content'].strip()

    # เขียนสรุปผลที่ได้ลงใน Sheet1
    sheet1 = workbook['Sheet1']
    sheet1.cell(row=1, column=1, value="Summary")
    sheet1.cell(row=2, column=1, value=summary)

    # บันทึกผลลัพธ์ลงไฟล์ Excel ใหม่
    result_path = DOWNLOAD_FILE_NAME
    workbook.save(result_path)

    # เพิ่มกราฟไปยังไฟล์ Excel
    create_graph_and_export(data, interval, result_path)

    print(f"Summary and graph added to {result_path}")


@app.route('/api/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return abort(400, description="No file part")
    
    file = request.files['file']
    if file.filename == '':
        return abort(400, description="No selected file")
    
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(file_path)
        
        # กำหนดค่า interval ที่เหมาะสมที่นี่
        interval = 10  # ตัวอย่างค่า

        # ประมวลผลไฟล์ที่อัปโหลด
        process_file(file_path, interval)

        return jsonify({"message": "File processed successfully"}), 200
    else:
        return abort(400, description="Invalid file type")


@app.route('/api/download', methods=['GET'])
def download_file(): 
    # หา path ของ root directory และไฟล์ที่ต้องการดาวน์โหลด
    root_path = os.path.abspath(os.path.dirname(__file__))
    file_path = os.path.join(root_path, DOWNLOAD_FILE_NAME)

    if os.path.exists(file_path):
        return send_file(file_path, as_attachment=True)
    else:
        abort(404)  # ถ้าไม่พบไฟล์ให้คืนค่า 404

if __name__ == "__main__":
    app.run(port=5000, debug=True)
