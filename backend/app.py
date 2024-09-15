from flask import Flask, request, jsonify, send_file, abort
import os
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment
import openai

from dotenv import load_dotenv
import os
import openai

# โหลดไฟล์ .env
load_dotenv()

# ดึงค่า OPENAI_API_KEY จาก .env
openai.api_key = os.getenv('OPENAI_API_KEY')

# ใช้งาน openai API ตามปกติ



app = Flask(__name__)

# ตั้งค่า OpenAI API key ของคุณ
openai.api_key = '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'  # เปลี่ยนเป็น API key ที่ถูกต้อง

# กำหนดที่เก็บไฟล์
UPLOAD_FOLDER = 'uploads'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# ชื่อไฟล์ที่ต้องการดาวน์โหลด
DOWNLOAD_FILE_NAME = 'result_with_summary.xlsx'

# ฟังก์ชันประมวลผลไฟล์
def process_file(file_path):
    try:
        data = pd.read_csv(file_path)
    except Exception as e:
        print(f"Error: Unable to read CSV file. {e}")
        abort(400, description="Unable to read CSV file.")

    # ตรวจสอบคอลัมน์ที่มีอยู่
    print("Columns in the DataFrame:", data.columns)

    # ตรวจสอบว่ามีคอลัมน์ 'label' อยู่ใน DataFrame หรือไม่
    label_column = 'label'
    if label_column not in data.columns:
        print("Error: 'label' column not found in the data!")
        abort(400, description="'label' column not found in the data!")

    # ตรวจสอบค่าในคอลัมน์ 'success'
    print("Unique values in 'success' column:", data['success'].unique())

    # แปลงค่า `elapsed` จาก ms เป็น s
    data['elapsed_seconds'] = data['elapsed'] / 1000

    # คำนวณค่า Pass, Fail และ %Error Rate
    data['Pass'] = data['success'].apply(lambda x: 1 if x else 0)
    data['Fail'] = data['success'].apply(lambda x: 0 if x else 1)

    # รวมค่าเส้น API ที่มีชื่อเดียวกัน
    summarized_data = data.groupby('label').agg(
        Transaction_Name=('label', 'first'),
        Expected=('label', 'size'),
        Pass=('Pass', 'sum'),
        Fail=('Fail', 'sum'),
        AvgResponseTime=('elapsed_seconds', 'mean'),
        MinResponseTime=('elapsed_seconds', 'min'),
        MaxResponseTime=('elapsed_seconds', 'max')
    ).reset_index(drop=True)

    # คำนวณ %Error Rate หลังจากการรวมข้อมูล
    summarized_data['%Error Rate'] = (summarized_data['Fail'] / summarized_data['Expected']) * 100

    # กำหนดค่า SLA ของ Response Time ให้เป็น 3 ทั้งหมด
    summarized_data['SLA_ResponseTime'] = 3

    # สร้างคอลัมน์ Status สำหรับการตรวจสอบการผ่านหรือไม่ผ่าน
    summarized_data['Status'] = summarized_data['%Error Rate'].apply(lambda x: "Pass" if x < 2 else "Failed")

    # โหลดเทมเพลต Excel
    template_path = "template.xlsx"
    workbook = openpyxl.load_workbook(template_path)
    sheet2 = workbook['Sheet2']

    # กำหนดสีสำหรับสถานะ
    green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

    # กำหนดกรอบรอบบรรทัดที่มีข้อมูล
    thin_border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))

    # กำหนดการจัดตำแหน่งชิดขวาสำหรับข้อมูลตัวเลข
    right_align = Alignment(horizontal='right')

    # เขียนข้อมูลที่รวมค่าแล้วลง Excel (Sheet2)
    for index, row in summarized_data.iterrows():
        start_row = index + 3  # สามารถปรับแถวเริ่มต้นตามที่ต้องการ

        sheet2.cell(row=start_row, column=1, value=row['Transaction_Name']).border = thin_border

        cell_expected = sheet2.cell(row=start_row, column=2, value=row['Expected'])
        cell_expected.border = thin_border
        cell_expected.alignment = right_align

        cell_pass = sheet2.cell(row=start_row, column=3, value=row['Pass'])
        cell_pass.border = thin_border
        cell_pass.alignment = right_align

        cell_fail = sheet2.cell(row=start_row, column=4, value=row['Fail'])
        cell_fail.border = thin_border
        cell_fail.alignment = right_align

        cell_error_rate = sheet2.cell(row=start_row, column=5, value=f"{row['%Error Rate']:.2f}%")
        cell_error_rate.border = thin_border
        cell_error_rate.alignment = right_align

        status_cell = sheet2.cell(row=start_row, column=6, value=row['Status'])
        if row['%Error Rate'] < 2:
            status_cell.fill = green_fill
        else:
            status_cell.fill = red_fill
        status_cell.border = thin_border

        cell_sla_response_time = sheet2.cell(row=start_row, column=7, value=row['SLA_ResponseTime'])
        cell_sla_response_time.border = thin_border
        cell_sla_response_time.alignment = right_align

        cell_avg_response_time = sheet2.cell(row=start_row, column=8, value=row['AvgResponseTime'])
        cell_avg_response_time.border = thin_border
        cell_avg_response_time.alignment = right_align

        cell_min_response_time = sheet2.cell(row=start_row, column=9, value=row['MinResponseTime'])
        cell_min_response_time.border = thin_border
        cell_min_response_time.alignment = right_align

        cell_max_response_time = sheet2.cell(row=start_row, column=10, value=row['MaxResponseTime'])
        cell_max_response_time.border = thin_border
        cell_max_response_time.alignment = right_align

        # เปรียบเทียบ SLA ของ Response Time กับ Average
        sla_status_cell = sheet2.cell(row=start_row, column=11)
        if row['AvgResponseTime'] <= row['SLA_ResponseTime']:
            sla_status_cell.value = "Pass"
            sla_status_cell.fill = green_fill
        else:
            sla_status_cell.value = "Failed"
            sla_status_cell.fill = red_fill
        sla_status_cell.border = thin_border
        sla_status_cell.alignment = right_align

    # **การสรุปผลด้วย OpenAI API**

    # อ่านข้อมูลจาก Sheet2 เข้า DataFrame อีกครั้ง
    sheet2_data = pd.DataFrame(sheet2.values)
    sheet2_data.columns = sheet2_data.iloc[0]  # ตั้งค่าแถวที่ 0 เป็นชื่อคอลัมน์
    sheet2_data = sheet2_data[1:]  # เอาข้อมูลจริงๆ ที่เริ่มจากแถวที่ 1 เป็นต้นไป

    # รวมข้อมูลเพิ่มเติมที่ต้องการ
    test_type = "Load Test"  # ตัวอย่างของประเภทที่ใช้ในการทดสอบ
    total_users = data['threadName'].nunique()  # จำนวนนักทดสอบ (user) ทั้งหมด
    project_name = "Example Project"  # ตัวอย่างของชื่อโปรเจค

    # แปลง DataFrame เป็นข้อความเพื่อส่งไปยัง OpenAI
    text_data = sheet2_data.to_string(index=False)

    # โปรมต์ที่ต้องการให้ OpenAI สรุปข้อมูล
    prompt = (f"Please summarize the following data:\n\n{text_data}\n\n"
            f"Test Type: {test_type}\n"
            f"Total Users: {total_users}\n"
            f"Project Name: {project_name}\n\n"
            f"Detail:\n"
            f"- Show all Passed , SLA transactions\n"
            f"- Show all Failed , SLA transactions\n\n"
            f"Summary:")

    # เรียกใช้ OpenAI API เพื่อให้สรุปข้อมูล
    response = openai.ChatCompletion.create(
        model="gpt-4",
        messages=[
            {"role": "system", "content": "You are a data analyst."},
            {"role": "user", "content": prompt}
        ]
    )

    summary = response.choices[0].message['content'].strip()

    # เขียนสรุปผลที่ได้ลงใน Sheet1
    sheet1 = workbook['Sheet1']
    sheet1.cell(row=1, column=1, value="Summary")
    sheet1.cell(row=2, column=1, value=summary)

    # บันทึกผลลัพธ์ลงไฟล์ Excel ใหม่
    result_path = DOWNLOAD_FILE_NAME
    workbook.save(result_path)

    print(f"Summary added to {result_path}")

@app.route('/api/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({"error": "No file part"}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({"error": "No selected file"}), 400

    if file:
        # หา path ของ root ของโปรเจค
        root_path = os.path.abspath(os.path.dirname(__file__))

        # กำหนด path ของไฟล์ที่จะบันทึกใน root directory
        file_path = os.path.join(root_path, file.filename)
        file.save(file_path)

        # ประมวลผลไฟล์ที่อัปโหลด
        process_file(file_path)

        # ส่ง response กลับไปบอกว่าการประมวลผลเสร็จสิ้นแล้ว
        return jsonify({"message": "File uploaded and processed successfully."}), 200

    
@app.route('/api/download', methods=['GET'])
def download_file(): 
    # หา path ของ root directory และไฟล์ที่ต้องการดาวน์โหลด
    root_path = os.path.abspath(os.path.dirname(__file__))
    file_path = os.path.join(root_path, DOWNLOAD_FILE_NAME)

    if os.path.exists(file_path):
        return send_file(file_path, as_attachment=True)
    else:
        abort(404)  # ถ้าไม่พบไฟล์ให้คืนค่า 404


if __name__ == "__main__":
    app.run(port=5000, debug=True)
